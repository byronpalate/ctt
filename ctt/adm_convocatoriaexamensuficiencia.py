# coding=utf-8
from datetime import datetime
import openpyxl
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponseRedirect
from django.http.response import HttpResponseServerError
from django.shortcuts import render

from decorators import secure_module, last_access
from settings import ARCHIVO_TIPO_PUBLICO, SOLICITUD_NUMERO_AUTOMATICO, TIPO_IVA_0_ID, \
    RUBRO_OTRO_SOLICITUD_ID, RUBRO_OTRO_EXAMEN_UBICACION_ID
from ctt.commonviews import adduserdata, obtener_reporte
from ctt.forms import ProcesoExamenSuficienciaForm, DetalleConvocatoriaExamenSuficienciaForm, \
    RequisitosDetalleFechaProcesoExamenSuficienciaForm, RangosNotasExamenInglesForm, \
    ArchivoListadoAprobadosExamenComplexivoForm, ImportarArchivoXLSPeriodoForm, FechaExamenSuficienciaForm, \
    RegistroEstudianteExamenInglesForm
from ctt.funciones import MiPaginador, log, bad_json, ok_json, url_back, generar_nombre, \
    fechatope_examenubicacion_ingles
from ctt.models import ConvocatoriaExamenSuficiencia, DetalleConvocatoriaExamenSuficiencia, \
    ProcesoAplicanteExamenSuficiencia, \
    RequisitosDetalleConvocatoriaExamenSuficiencia, null_to_numeric, RangosNotasExamenUbicacionIngles, \
    RangosNotasExamenValidacionIngles, Archivo, AsignaturaMalla,  Asignatura, PreValidacionIngles, Persona, \
    TipoSolicitudSecretariaDocente, SolicitudSecretariaDocente, HistorialSolicitud, Rubro, RubroOtro, \
    Periodo, RequisitosProcesoAplicanteSuficiencia, Inscripcion
from ctt.utils.etools.certificacionb1ingles import tipo_convocatoria_b1
from ctt.tasks import send_mail


def inscripciones_para_convocatoria(convocatoria):
    inscripciones = Inscripcion.objects.filter(sede=convocatoria.sede)

    if convocatoria.coordinacion.exists():
        inscripciones = inscripciones.filter(coordinacion__in=convocatoria.coordinacion.all())

    if convocatoria.modalidad.exists():
        inscripciones = inscripciones.filter(modalidad__in=convocatoria.modalidad.all())

    return inscripciones.exclude(carrera_id=113).distinct()


def validar_inscripcion_convocatoria(inscripcion, convocatoria):
    if convocatoria.coordinacion.exists() and not convocatoria.coordinacion.filter(id=inscripcion.coordinacion_id).exists():
        return u'El estudiante no pertenece a la facultad designada para esta convocatoria.'

    if convocatoria.modalidad.exists() and not convocatoria.modalidad.filter(id=inscripcion.modalidad_id).exists():
        return u'El estudiante no pertenece a la modalidad designada para esta convocatoria.'

    if convocatoria.sede_id and inscripcion.sede_id != convocatoria.sede_id:
        return u'El estudiante no pertenece a la sede designada para esta convocatoria.'

    return None


def crear_registro_examen_ingles(inscripcion, convocatoria, request, generar_rubro=True):
    if convocatoria.registro(inscripcion):
        return None

    hoy = datetime.now().date()
    registro = ProcesoAplicanteExamenSuficiencia(
        convocatoria=convocatoria,
        inscripcion=inscripcion,
        fechaaplicacion=datetime.today(),
        promedionotas=0,
        fechatope=fechatope_examenubicacion_ingles(hoy, inscripcion)
    )
    registro.save(request)

    for actividad in convocatoria.detalleconvocatoriaexamensuficiencia_set.all():
        for requisito in actividad.requisitosdetalleconvocatoriaexamensuficiencia_set.all():
            RequisitosProcesoAplicanteSuficiencia(proceso=registro, requisito=requisito).save(request)

    if generar_rubro:
        tipo = TipoSolicitudSecretariaDocente.objects.get(pk=57)
        valor = null_to_numeric(tipo.valor + tipo.costo_base, 2)
        rubro = Rubro(
            inscripcion=inscripcion,
            valor=valor,
            iva_id=TIPO_IVA_0_ID,
            valortotal=valor,
            saldo=valor,
            periodo=convocatoria.periodo,
            fecha=hoy,
            fechavence=hoy
        )
        rubro.save(request)
        RubroOtro(rubro=rubro, tipo_id=RUBRO_OTRO_EXAMEN_UBICACION_ID).save(request)
        rubro.actulizar_nombre('EXAMEN DE ' + convocatoria.tipoconvocatoria.nombrerubro + ' DE INGLES')

    return registro


@login_required(login_url='/login')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    adduserdata(request, data)
    sedeseleccionada = request.session['coordinacionseleccionada'].sede
    coordinacion = request.session['coordinacionseleccionada']
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                form = ProcesoExamenSuficienciaForm(request.POST)
                if form.is_valid():
                    if not form.cleaned_data['modalidad']:
                        return bad_json(mensaje=u'Debe escoger al menos una carrera')
                    proceso = ConvocatoriaExamenSuficiencia(nombre=form.cleaned_data['nombre'],
                                                            tipoconvocatoria=form.cleaned_data['tipoconvocatoria'],
                                                            fechainicio=form.cleaned_data['fechainicio'],
                                                            fechafin=form.cleaned_data['fechafin'],
                                                            periodo=request.session['periodo'],
                                                            autoregistro=form.cleaned_data['autoregistro'],
                                                            activo=True if form.cleaned_data['autoregistro'] else False,
                                                            sede=sedeseleccionada,
                                                            mensaje=form.cleaned_data['mensaje'],
                                                            fechainicioautoregistro=form.cleaned_data['fechainicioautoregistro'],
                                                            fechafinautoregistro=form.cleaned_data['fechafinautoregistro'])
                    proceso.save(request)
                    proceso.modalidad.set(form.cleaned_data['modalidad'])
                    proceso.coordinacion.set(form.cleaned_data['coordinacion'])
                    log(u'Adicionado proceso examen complexivo: %s' % proceso, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'edit':
            try:
                form = ProcesoExamenSuficienciaForm(request.POST)
                if form.is_valid():
                    if not form.cleaned_data['modalidad']:
                        return bad_json(mensaje=u'Debe escoger al menos una modalidad')
                    proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    proceso.nombre = form.cleaned_data['nombre']
                    proceso.tipoconvocatoria = form.cleaned_data['tipoconvocatoria']
                    proceso.fechainicio = form.cleaned_data['fechainicio']
                    proceso.fechafin = form.cleaned_data['fechafin']
                    proceso.mensaje = form.cleaned_data['mensaje']
                    proceso.autoregistro = form.cleaned_data['autoregistro']
                    proceso.fechainicioautoregistro = form.cleaned_data['fechainicioautoregistro']
                    proceso.fechafinautoregistro = form.cleaned_data['fechafinautoregistro']
                    proceso.save(request)
                    proceso.modalidad.set(form.cleaned_data['modalidad'])
                    proceso.coordinacion.set(form.cleaned_data['coordinacion'])
                    log(u'Edito convocatoria a examen: %s' % proceso, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'editfechaexamen':
            try:
                form = FechaExamenSuficienciaForm(request.POST)
                if form.is_valid():
                    proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    proceso.fechaexamen = form.cleaned_data['fecha']
                    proceso.save(request)
                    log(u'Edito la fecha de rendicion de examen de ingles: %s' % proceso, request, "edit")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'habilitar':
            try:
                detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                detalle.activo = True
                detalle.save(request)
                log(u'Activo proceso consultorio juridico: %s' % detalle, request, "add")
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'autoregistrar':
            try:
                convocatoria = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                for inscripcion in inscripciones_para_convocatoria(convocatoria):
                    registro = crear_registro_examen_ingles(
                        inscripcion,
                        convocatoria,
                        request,
                        generar_rubro=inscripcion.modalidad_id not in (1, 5)
                    )
                    if not registro:
                        continue
                    log(u'Se auto registra a: %s en la convocatoria: %s' % (inscripcion, convocatoria), request, "add")
                    if not inscripcion.carrera.posgrado:
                        send_mail(subject=('Registro en Examén de Ubicación'),
                                  html_template='emails/avisoautoregistro.html',
                                  data={'proceso': convocatoria, 'persona': registro.inscripcion.persona},
                                  recipient_list=[registro.inscripcion.persona])
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'deshabilitar':
            try:
                detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                detalle.activo = False
                detalle.save(request)
                log(u'Deshabilito Procesos de convocatoria al examen de suficiencia: %s' % detalle, request, "edit")
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'del':
            try:
                convocatoria = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                log(u'Elimino proceso examen suficiencia: %s' % convocatoria, request, "del")
                convocatoria.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'adddetalle':
            try:
                convocatoria = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                form = DetalleConvocatoriaExamenSuficienciaForm(request.POST)
                if form.is_valid():
                    inicio = form.cleaned_data['inicio']
                    fin = form.cleaned_data['fin']
                    if fin < inicio:
                        return bad_json(mensaje=u'La fecha fin no puede ser menor a la fecha inicio.')
                    detalle = DetalleConvocatoriaExamenSuficiencia(convocatoriaconsultorio=convocatoria,
                                                                   descripcion=form.cleaned_data['descripcion'],
                                                                   inicio=form.cleaned_data['inicio'],
                                                                   fin=form.cleaned_data['fin'],
                                                                   informativo=form.cleaned_data['informativo'])
                    detalle.save(request)
                    log(u'Adicionado detalle proceso examen complexivo: %s' % detalle, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'editdetalle':
            try:
                form = DetalleConvocatoriaExamenSuficienciaForm(request.POST)
                if form.is_valid():
                    detalle = DetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    detalle.descripcion = form.cleaned_data['descripcion']
                    detalle.inicio = form.cleaned_data['inicio']
                    detalle.fin = form.cleaned_data['fin']
                    detalle.informativo = form.cleaned_data['informativo']
                    detalle.save(request)
                    log(u'Adicionado proceso examen suficiencia: %s' % detalle, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'deldetalle':
            try:
                detalle = DetalleConvocatoriaExamenSuficiencia.objects.get(id=request.POST['id'])
                log(u'Elimino detalle: %s' % detalle, request, "del")
                detalle.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'addrequisitoexamen':
            try:
                form = RequisitosDetalleFechaProcesoExamenSuficienciaForm(request.POST)
                detalle = DetalleConvocatoriaExamenSuficiencia.objects.get(id=request.POST['id'])
                if form.is_valid():
                    requisito = RequisitosDetalleConvocatoriaExamenSuficiencia(detalleproceso=detalle,
                                                                               tipo=form.cleaned_data['tipo'],
                                                                               obligatorio=form.cleaned_data['obligatorio'])
                    requisito.save(request)
                    log(u'Adicionado requisito examen suficiencia: %s' % requisito, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'editrequisitoexamen':
            try:
                form = RequisitosDetalleFechaProcesoExamenSuficienciaForm(request.POST)
                requisito = RequisitosDetalleConvocatoriaExamenSuficiencia.objects.get(id=request.POST['id'])
                if form.is_valid():
                    requisito.tipo=form.cleaned_data['tipo']
                    requisito.obligatorio=form.cleaned_data['obligatorio']
                    requisito.save(request)
                    log(u'Edita requisito de consultorio: %s' % requisito, request, "edit")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'delrequisitoexamen':
            try:
                requisito = RequisitosDetalleConvocatoriaExamenSuficiencia.objects.get(id=request.POST['id'])
                log(u'Elimino requisito de consultorio: %s' % requisito, request, "del")
                requisito.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'addlistado':
            try:
                form = ArchivoListadoAprobadosExamenComplexivoForm(request.POST, request.FILES)
                if form.is_valid():
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("listadoaprobados", newfile._name)
                    convocatoria = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    convocatoria.listadoaprobados = newfile
                    convocatoria.save(request)
                    log(u'Adiciono archivo: %s' % convocatoria, request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'calcularnotafinal':
            try:
                registro = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                nota = float(request.POST['nota'])
                promedio = float(request.POST['promedio'])
                notafinal = null_to_numeric((nota * 0.4)+(promedio * 0.6),2)
                registro.notafinal = notafinal
                registro.notaexamenconocimientos = nota
                registro.save()
                log(u'Calculo nota final: %s - %s - %s' % (registro.notafinal,registro.notaexamenconocimientos,registro.inscripcion), request, "add")
                return ok_json({'notafinal': notafinal})
            except:
                return bad_json(error=3)

        if action == 'calcularnotafinal':
            try:
                registro = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                nota = float(request.POST['nota'])
                promedio = float(request.POST['promedio'])
                notafinal = null_to_numeric((nota * 0.4)+(promedio * 0.6),2)
                registro.notafinal = notafinal
                registro.notaexamenconocimientos = nota
                registro.save()
                log(u'Calculo nota final: %s - %s - %s' % (registro.notafinal,registro.notaexamenconocimientos,registro.inscripcion), request, "add")
                return ok_json({'notafinal': notafinal})
            except:
                return bad_json(error=3)

        if action == 'calcularmodulos':
            try:
                registrado=ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                proceso=registrado.convocatoria
                if proceso.tipoconvocatoria_id == 1:  # ESTO ES PARA VALIDACION
                    rangos = RangosNotasExamenValidacionIngles.objects.all().order_by('nota')
                else:
                    rangos = RangosNotasExamenUbicacionIngles.objects.all().order_by('nivel')
                if not rangos.exists():
                    return bad_json(mensaje=u"No tiene rangos de puntaje asignados.")
                persona = registrado.inscripcion.persona
                puntejeobtenido =float(request.POST.get('nota', 0))
                nota=0
                if proceso.tipoconvocatoria_id == 1:  # ESTO ES PARA VALIDACION
                    for rango in rangos:
                        if puntejeobtenido >= rango.inicio and puntejeobtenido <= rango.fin:
                            entro = True
                            nota = rango.nota
                            PreValidacionIngles.objects.filter(proceso=registrado).delete()
                            if rango.aprueba == True:
                                malla = registrado.inscripcion.mi_malla()
                                # preguntar si se quedan con la misma nota o se cambia
                                aprobadasmodulos = AsignaturaMalla.objects.filter(recordacademico__inscripcion=registrado.inscripcion, recordacademico__aprobada=True).values_list('id', flat=True)
                                inglespendientes = AsignaturaMalla.objects.filter(malla=malla).exclude(id__in=aprobadasmodulos).order_by('asignatura')
                                for materia in inglespendientes:
                                    asignatura = Asignatura.objects.get(pk=materia.asignatura_id)
                                    PreValidacionIngles.objects.create(proceso=registrado,
                                                                       asignatura=asignatura,
                                                                       asignaturamalla=materia,
                                                                       periodo=registrado.convocatoria.periodo,
                                                                       fecha=datetime.now(),
                                                                       nota_conocimiento=nota,
                                                                       nota_final=nota,
                                                                       estadocon=True,
                                                                       horas=materia.horas)
                                break
                    registrado.notaexamenconocimientos = puntejeobtenido
                    registrado.notafinal = nota
                    registrado.save()
                    log(u'Modifico la nota de examen de validacion del estudiante %s a la nota %s' % (registrado.inscripcion.persona, nota), request, "add")
                    return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'calcularmodulosubi':
            try:
                registrado=ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                proceso=registrado.convocatoria
                rangos = RangosNotasExamenUbicacionIngles.objects.all().order_by('nivel')
                if not rangos.exists():
                    return bad_json(mensaje=u"No tiene rangos de puntaje asignados.")
                persona = registrado.inscripcion.persona
                puntejeobtenido =float(request.POST.get('nota', 0))
                nota=0
                # Definir niveles en un diccionario
                niveles_dict = {
                    0: [0],
                    1: [4311],
                    2: [4311, 4312],
                    3: [4311, 4312, 4313],
                    4: [4311, 4312, 4313, 4316],
                    5: [4311, 4312, 4313, 4316, 4317],
                    6: [4311, 4312, 4313, 4316, 4317, 4318]  # Puedes agregar más niveles si es necesario
                }
                for rango in rangos:
                    if puntejeobtenido >= rango.inicio and puntejeobtenido <= rango.fin:
                        entro = True
                        nota = 10
                        PreValidacionIngles.objects.filter(proceso=registrado).delete()

                        malla = registrado.inscripcion.mi_malla()
                        # preguntar si se quedan con la misma nota o se cambia
                        aprobadasmodulos = AsignaturaMalla.objects.filter(
                            recordacademico__inscripcion=registrado.inscripcion,
                            recordacademico__aprobada=True).values_list('id', flat=True)
                        # Obtener niveles según rango.nivel, si no está en el diccionario, toma el máximo
                        niveles = niveles_dict.get(rango.nivel)
                        inglespendientes = AsignaturaMalla.objects.filter(malla=malla, asignatura_id__in=niveles).exclude(
                            id__in=aprobadasmodulos).order_by('asignatura')
                        for materia in inglespendientes:
                            asignatura = Asignatura.objects.get(pk=materia.asignatura_id)
                            PreValidacionIngles.objects.create(
                                proceso=registrado,
                                asignatura=asignatura,
                                asignaturamalla=materia,
                                periodo=registrado.convocatoria.periodo,
                                fecha=datetime.now(),
                                nota_conocimiento=nota,
                                nota_final=nota,
                                estadocon=True,
                                horas=materia.horas
                            )
                        break
                    else:
                        PreValidacionIngles.objects.filter(proceso=registrado).delete()
                registrado.notaexamenconocimientos = puntejeobtenido
                registrado.notafinal = nota
                registrado.save()
                log(u'Modifico la nota de examen de validacion del estudiante %s a la nota %s' % (registrado.inscripcion.persona, nota),
                    request, "add")
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'delregistro':
            try:
                registro = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                log(u'Elimino registro: %s' % registro, request, "del")
                registro.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'addrangoubicacion':
            try:
                form = RangosNotasExamenInglesForm(request.POST)
                if form.is_valid():
                    rango = RangosNotasExamenUbicacionIngles(inicio=form.cleaned_data['inicio'],
                                                             fin=form.cleaned_data['fin'],
                                                             nivel=form.cleaned_data['nivel'])
                    rango.save(request)
                    log(u'Adicionado un rango de notas para examen de ubicacion de %s a %s' % (rango.inicio, rango.fin), request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'editrangonota':
            try:
                form = RangosNotasExamenInglesForm(request.POST)
                rango = RangosNotasExamenUbicacionIngles.objects.get(id=request.POST['id'])
                if form.is_valid():
                    rango.inicio=form.cleaned_data['inicio']
                    rango.fin=form.cleaned_data['fin']
                    rango.nivel=form.cleaned_data['nivel']
                    rango.save(request)
                    log(u'Edita el rango de puntacion para examen de ubicacion: %s' % rango, request, "edit")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'delrango':
            try:
                rango = RangosNotasExamenUbicacionIngles.objects.get(pk=request.POST['id'])
                log(u'Elimino el rango de examen de puntuacion: %s' % rango, request, "edit")
                rango.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'addrangovalidacion':
            try:
                form = RangosNotasExamenInglesForm(request.POST)
                if form.is_valid():
                    rango = RangosNotasExamenValidacionIngles(inicio=form.cleaned_data['inicio'],
                                                              fin=form.cleaned_data['fin'],
                                                              aprueba=form.cleaned_data['aprueba'],
                                                              nota=form.cleaned_data['nota'])
                    rango.save(request)
                    log(u'Adicionado un rango de notas para examen de validacion de %s a %s' % (rango.inicio, rango.fin), request, "add")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'editrangonotaval':
            try:
                form = RangosNotasExamenInglesForm(request.POST)
                rango = RangosNotasExamenValidacionIngles.objects.get(id=request.POST['id'])
                if form.is_valid():
                    rango.inicio=form.cleaned_data['inicio']
                    rango.fin=form.cleaned_data['fin']
                    rango.aprueba=form.cleaned_data['aprueba']
                    rango.nota=form.cleaned_data['nota']
                    rango.save(request)
                    log(u'Edita el rango de puntacion para examen de validacion del ingles: %s' % rango, request, "edit")
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'delrangoval':
            try:
                rango = RangosNotasExamenValidacionIngles.objects.get(pk=request.POST['id'])
                log(u'Elimino el rango de examen de validacion de ingles: %s' % rango, request, "edit")
                rango.delete()
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'subirnotas':
            try:
                # DepurarAdmisiones.objects.all().delete()
                form = ImportarArchivoXLSPeriodoForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("notasexamenvalidacioningles", nfile._name)
                    archivo = Archivo(nombre='notasexamenvalidacioningles',
                                      fecha=datetime.now(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_PUBLICO)
                    archivo.save(request)
                    workbook = openpyxl.load_workbook(archivo.archivo.file.name)
                    sheet = workbook.worksheets[0]
                    linea = 1
                    data['proceso'] = proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    registrados = proceso.procesoaplicanteexamensuficiencia_set.all()
                    print(datetime.now().time())
                    if proceso.tipoconvocatoria_id == 1:  # ESTO ES PARA VALIDACION
                        rangos = RangosNotasExamenValidacionIngles.objects.all().order_by('nota')
                    else:
                        rangos=RangosNotasExamenUbicacionIngles.objects.all().order_by('nivel')
                    if not rangos.exists():
                        return bad_json(mensaje=u"No tiene rangos de puntaje asignados.")
                    for rowx in sheet.iter_rows(values_only=True):
                        if linea >= 2:
                            # en cols se almacena toda la tupla que trae de excel
                            cols = rowx
                            for registrado in registrados:
                                persona=registrado.inscripcion.persona
                                puntejeobtenido=cols[3]
                                if proceso.tipoconvocatoria_id == 1: #ESTO ES PARA VALIDACION
                                    if persona.cedula == str(cols[0]):
                                        for rango in rangos:
                                            if puntejeobtenido >= rango.inicio and puntejeobtenido <= rango.fin:
                                                entro = True
                                                nota=rango.nota
                                                PreValidacionIngles.objects.filter(proceso=registrado).delete()
                                                if rango.aprueba == True:
                                                    malla = registrado.inscripcion.mi_malla()
                                                    # preguntar si se quedan con la misma nota o se cambia
                                                    aprobadasmodulos = AsignaturaMalla.objects.filter(recordacademico__inscripcion=registrado.inscripcion,
                                                                                                  recordacademico__aprobada=True).values_list('id', flat=True)
                                                    inglespendientes = AsignaturaMalla.objects.filter(malla=malla).exclude(id__in=aprobadasmodulos).order_by('asignatura')
                                                    for materia in inglespendientes:
                                                        asignatura = Asignatura.objects.get(pk=materia.asignatura_id)
                                                        PreValidacionIngles.objects.create(proceso=registrado,
                                                                                           asignatura=asignatura,
                                                                                           asignaturamalla=materia,
                                                                                           periodo=registrado.convocatoria.periodo,
                                                                                           fecha=datetime.now(),
                                                                                           nota_conocimiento=nota,
                                                                                           nota_final=nota,
                                                                                           estadocon=True,
                                                                                           horas=materia.horas)
                                                    break
                                        registrado.notaexamenconocimientos=cols[3]
                                        registrado.notafinal=nota
                                        registrado.save()
                                        break
                        linea += 1
                        print(linea)
                    print(datetime.now().time())
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'subirnotasubi':
            try:
                # DepurarAdmisiones.objects.all().delete()
                form = ImportarArchivoXLSPeriodoForm(request.POST, request.FILES)
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("notasexamenubicacioningles", nfile._name)
                    archivo = Archivo(nombre='notasexamenubicacioningles',
                                      fecha=datetime.now(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_PUBLICO)
                    archivo.save(request)
                    workbook = openpyxl.load_workbook(archivo.archivo.file.name)
                    sheet = workbook.worksheets[0]
                    linea = 1
                    data['proceso'] = proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['id'])
                    registrados = proceso.procesoaplicanteexamensuficiencia_set.all()
                    rangos=RangosNotasExamenUbicacionIngles.objects.all().order_by('nivel')
                    if not rangos.exists():
                        return bad_json(mensaje=u"No tiene rangos de puntaje asignados, para examén de Ubicación.")
                        # Definir niveles en un diccionario
                    niveles_dict = {
                        0: [0],
                        1: [4311],
                        2: [4311, 4312],
                        3: [4311, 4312, 4313],
                        4: [4311, 4312, 4313,4316],
                        5: [4311, 4312, 4313,4316,4317],
                        6: [4311, 4312, 4313,4316,4317,4318]  # Puedes agregar más niveles si es necesario
                    }
                    for rowx in sheet.iter_rows(values_only=True):
                        if linea >= 2:
                            # en cols se almacena toda la tupla que trae de excel
                            cols = rowx
                            nota=0
                            for registrado in registrados:
                                persona=registrado.inscripcion.persona
                                puntajeobtenido=cols[3]
                                if persona.cedula == str(cols[0]):
                                    for rango in rangos:
                                        if puntajeobtenido >= rango.inicio and puntajeobtenido <= rango.fin:
                                            entro = True
                                            nota=10
                                            PreValidacionIngles.objects.filter(proceso=registrado).delete()
                                            malla = registrado.inscripcion.mi_malla()
                                            aprobadasmodulos = AsignaturaMalla.objects.filter(recordacademico__inscripcion=registrado.inscripcion,
                                                                                          recordacademico__aprobada=True).values_list('id', flat=True)
                                            # Obtener niveles según rango.nivel, si no está en el diccionario, toma el máximo
                                            niveles = niveles_dict.get(rango.nivel)
                                            inglespendientes = AsignaturaMalla.objects.filter(malla=malla,asignatura_id__in=niveles ).exclude(id__in=aprobadasmodulos).order_by('asignatura')
                                            for materia in inglespendientes:
                                                asignatura = Asignatura.objects.get(pk=materia.asignatura_id)
                                                PreValidacionIngles.objects.create(
                                                    proceso=registrado,
                                                    asignatura=asignatura,
                                                    asignaturamalla=materia,
                                                    periodo=registrado.convocatoria.periodo,
                                                    fecha=datetime.now(),
                                                    nota_conocimiento=nota,
                                                    nota_final=nota,
                                                    estadocon=True,
                                                    horas=materia.horas)
                                            break
                                        else:
                                            PreValidacionIngles.objects.filter(proceso=registrado).delete()
                                    registrado.notaexamenconocimientos = cols[3]
                                    registrado.notafinal = nota
                                    registrado.save()
                        linea += 1
                        print(linea)
                    print(datetime.now().time())
                    return ok_json()
                else:
                    return bad_json(error=6,form=form)
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        if action == 'pasarrecord':
            try:
                proceso = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                for modulos in PreValidacionIngles.objects.filter(proceso=proceso,estadofin=1):
                    modulos.pasar_record_modulo()
                    if proceso.convocatoria.tipoconvocatoria == 1:
                        log(u'Paso a record de %s ,por examen de validacion el modulos de:  %s' % (proceso.inscripcion.persona,modulos.asignatura), request, "add")
                    else:
                        log(u'Paso a record de %s ,por examen de ubicacion el modulos de:  %s' % (proceso.inscripcion.persona,modulos.asignatura), request, "add")
                proceso.pasorecord=True
                proceso.fecha_pasarecord = datetime.now()
                proceso.save()
                if proceso.convocatoria.tipoconvocatoria.id == 2:
                    # persona = Persona.objects.get(cedula='1804945291')
                    send_mail(subject=('Actualización Record Académico'),
                              html_template='emails/avisopasomoduloinglesrecord.html',
                              data={'proceso': proceso,
                                    'modulos': PreValidacionIngles.objects.filter(proceso=proceso, estadofin=1)
                                    },
                              # recipient_list=[persona])
                              recipient_list=[proceso.inscripcion.persona])
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                pass

        if action == 'generarrubrosvalidacion':
            try:
                proceso = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.POST['id'])
                if proceso.generorubros:
                    return bad_json('Los rubros ya fueron generados de manera automatica, en caso de requerir volver a generarlos los tiene que realizar el estudiante.')
                valorpagar=0
                for modulo in PreValidacionIngles.objects.filter(proceso=proceso, estadofin=1):
                    #RESPONSABLE DE SOLICITUDES DE INGLES CRISTINA AMANCHA
                    responsable = Persona.objects.get(id=42379)
                    #SOLICITUD DE APROBACION DE NIVEL DE IDIOMAS
                    tipo = TipoSolicitudSecretariaDocente.objects.get(id=12)
                    solicitud = SolicitudSecretariaDocente(fecha=datetime.now().date(),
                                                           hora=datetime.now().time(),
                                                           inscripcion=proceso.inscripcion,
                                                           tipo=tipo,
                                                           descripcion='APROBACION DEL MODULO DE INGLES POR EXAMEN DE VALIDACIÓN: ' + str(modulo.asignatura),
                                                           cerrada=False,
                                                           responsable=responsable,
                                                           matricula=None)
                    solicitud.save(request)
                    if responsable:
                        historial = HistorialSolicitud(solicitud=solicitud,
                                                       fecha=datetime.now(),
                                                       persona=solicitud.responsable,
                                                       respuesta='')
                        historial.save(request)
                    if SOLICITUD_NUMERO_AUTOMATICO:
                        if SolicitudSecretariaDocente.objects.filter(numero_tramite__gt=0).exists():
                            ultima = SolicitudSecretariaDocente.objects.filter(numero_tramite__gt=0).order_by('-id')[0]
                            solicitud.numero_tramite = ultima.numero_tramite + 1
                        else:
                            solicitud.numero_tramite = 1
                        solicitud.save(request)
                    if solicitud.tipo.tiene_costo():
                        cantidad = 1
                        if solicitud.tipo.costo_unico:
                            valor = null_to_numeric(solicitud.tipo.valor + solicitud.tipo.costo_base, 2)
                        else:
                            valor = null_to_numeric((solicitud.tipo.valor * cantidad) + solicitud.tipo.costo_base, 2)
                        # for p in PeriodoSolicitud.objects.all():
                        #     if p.vigente():
                        #         periodosolicitud = p.id
                        periodosolicitud=Periodo.objects.filter(parasolicitudes=True)[0]
                        valorpagar+=valor
                        rubro = Rubro(inscripcion=proceso.inscripcion,
                                      valor=valor,
                                      iva_id=TIPO_IVA_0_ID,
                                      valortotal=valor,
                                      saldo=valor,
                                      periodo=periodosolicitud,
                                      fecha=datetime.now().date(),
                                      fechavence=datetime.now().date())
                        rubro.save(request)
                        rubrootro = RubroOtro(rubro=rubro,
                                              tipo_id=RUBRO_OTRO_SOLICITUD_ID,
                                              solicitud=solicitud)
                        rubrootro.save(request)
                        rubro.actulizar_nombre(nombre=solicitud.tipo.nombre)
                        solicitud.verificar_gratuidad()
                        # solicitud.mail_subject_nuevo()
                        log(u'Se genera solicitud automatica para %s , por aprobacion en examen de validacion para el modulo:  %s' % (proceso.inscripcion.persona,modulo.asignatura), request, "add")
                proceso.generorubros=True
                proceso.save()
                send_mail(subject=('Valores generados examen de Validación'),
                          html_template='emails/valoresgeneradosingles.html',
                          data={'proceso': proceso,
                                'valor': valorpagar,
                                'modulos': PreValidacionIngles.objects.filter(proceso=proceso, estadofin=1)},
                          recipient_list=[proceso.inscripcion.persona])
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                pass

        if action == 'addregistro':
            try:
                proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.POST['convocatoria'])
                form = RegistroEstudianteExamenInglesForm(request.POST)
                if form.is_valid():
                    inscripcion = Inscripcion.objects.get(pk=form.cleaned_data['inscripcion'])
                    mensaje = validar_inscripcion_convocatoria(inscripcion, proceso)
                    if mensaje:
                        return bad_json(mensaje=mensaje)

                    registro = crear_registro_examen_ingles(
                        inscripcion,
                        proceso,
                        request,
                        generar_rubro=not proceso.autoregistro
                    )
                    if registro:
                        log(u'Se registro por parte de secretaria a : %s en la convocatoria: %s' % (registro.inscripcion, proceso), request, "add")
                else:
                    return bad_json(error=6, form=form)
                return ok_json()
            except Exception as ex:
                transaction.set_rollback(True)
                return bad_json(error=1, ex=ex)

        return bad_json(error=0)
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'add':
                try:
                    data['title'] = u'Agregar nuevo proceso'
                    data['form'] = ProcesoExamenSuficienciaForm(initial={'fechainicio': datetime.now().date(),
                                                                         'fechafin': datetime.now().date()})
                    return render(request, "adm_convocatoriaexamensuficiencia/add.html", data)
                except Exception as ex:
                    pass

            if action == 'edit':
                try:
                    data['title'] = u'Editar proceso'
                    data['proceso'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = ProcesoExamenSuficienciaForm(initial={'nombre': detalle.nombre,
                                                                         'tipoconvocatoria': detalle.tipoconvocatoria,
                                                                         'modalidad': detalle.modalidad.all(),
                                                                         'coordinacion': detalle.coordinacion.all(),
                                                                         'fechainicio': detalle.fechainicio,
                                                                         'fechafin': detalle.fechafin,
                                                                         'autoregistro': detalle.autoregistro,
                                                                         'mensaje': detalle.mensaje})
                    return render(request, "adm_convocatoriaexamensuficiencia/edit.html", data)
                except Exception as ex:
                    pass

            if action == 'habilitar':
                try:
                    data['title'] = u'Habilitar proceso'
                    data['proceso'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/habilitar.html", data)
                except Exception as ex:
                    pass

            if action == 'deshabilitar':
                try:
                    data['title'] = u'Deshabilitar proceso'
                    data['proceso'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/deshabilitar.html", data)
                except Exception as ex:
                    pass

            if action == 'del':
                try:
                    data['title'] = u'Eliminar convocatoria'
                    data['proceso'] = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delconvocatoria.html", data)
                except Exception as ex:
                    pass

            if action == 'detalles':
                try:
                    data['title'] = u'Cronograma'
                    data['proceso'] = proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['detalles'] = proceso.detalleconvocatoriaexamensuficiencia_set.all().order_by('inicio','fin')
                    return render(request, "adm_convocatoriaexamensuficiencia/detalles.html", data)
                except Exception as ex:
                    pass

            if action == 'registrados':
                try:
                    data['title'] = u'Registrados'
                    data['proceso'] = proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    registrados = proceso.procesoaplicanteexamensuficiencia_set.select_related("inscripcion__persona").order_by("inscripcion__persona__apellido1")
                    search = None
                    paging = MiPaginador(registrados, 30)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session and 'paginador_url' in request.session:
                            if request.session['paginador_url'] == 'adm_convocatoriaexamensuficiencia':
                                paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        page = paging.page(p)
                    except:
                        p = 1
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['paginador_url'] = 'adm_convocatoriaexamensuficiencia'
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['registrados'] = page.object_list
                    return render(request, "adm_convocatoriaexamensuficiencia/registrados.html", data)
                except Exception as ex:
                    pass

            if action == 'registradosubi':
                try:
                    data['title'] = u'Registrados Examén de Ubicación'
                    data['proceso'] = proceso = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    registrados = proceso.procesoaplicanteexamensuficiencia_set.select_related("inscripcion__persona").order_by("inscripcion__persona__apellido1")
                    search = None
                    paging = MiPaginador(registrados, 30)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session and 'paginador_url' in request.session:
                            if request.session['paginador_url'] == 'adm_convocatoriaexamensuficiencia':
                                paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        page = paging.page(p)
                    except:
                        p = 1
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['paginador_url'] = 'adm_convocatoriaexamensuficiencia'
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['registrados'] = page.object_list
                    return render(request, "adm_convocatoriaexamensuficiencia/registradosubi.html", data)
                except Exception as ex:
                    pass

            if action == 'adddetalle':
                try:
                    data['title'] = u'Cronograma'
                    data['proceso'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = DetalleConvocatoriaExamenSuficienciaForm()
                    return render(request, "adm_convocatoriaexamensuficiencia/adddetalle.html", data)
                except Exception as ex:
                    pass

            if action == 'editdetalle':
                try:
                    data['title'] = u'Editar cronograma'
                    data['detalle'] = detalle = DetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    form = DetalleConvocatoriaExamenSuficienciaForm(initial={'descripcion': detalle.descripcion,
                                                                             'inicio': detalle.inicio,
                                                                             'fin': detalle.fin,
                                                                             'informativo': detalle.informativo})
                    form.editar(detalle)
                    data['form'] = form
                    return render(request, "adm_convocatoriaexamensuficiencia/editdetalle.html", data)
                except Exception as ex:
                    pass

            if action == 'deldetalle':
                try:
                    data['title'] = u'Eliminar cronograma'
                    data['detalle'] = DetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/deldetalle.html", data)
                except Exception as ex:
                    pass

            if action == 'requisitosexamensuficiencia':
                try:
                    data['title'] = u'Requisitos para actividad examen suficiencia'
                    data['detalle'] = detalle = DetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['requisitos'] = detalle.requisitosdetalleconvocatoriaexamensuficiencia_set.all()
                    return render(request, "adm_convocatoriaexamensuficiencia/requisitosexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'revisardocumentacion':
                try:
                    data['proceso'] = proceso = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['p'] = request.GET['p']
                    data['title'] = u'Requisitos cargados de: ' + proceso.inscripcion.persona.nombre_completo()
                    data['requisitos'] = proceso.requisitosprocesoaplicantesuficiencia_set.all()
                    return render(request, "adm_convocatoriaexamensuficiencia/revisardocumentacion.html", data)
                except Exception as ex:
                    pass

            if action == 'addrequisitoexamen':
                try:
                    data['title'] = u'Adicionar requisitos examen suficiencia'
                    data['detalle'] = DetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = RequisitosDetalleFechaProcesoExamenSuficienciaForm()
                    return render(request, "adm_convocatoriaexamensuficiencia/addrequisitoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'editrequisitoexamen':
                try:
                    data['title'] = u'Editar requisito examen'
                    data['requisito'] = requisito = RequisitosDetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = form = RequisitosDetalleFechaProcesoExamenSuficienciaForm(initial={'tipo': requisito.tipo,
                                                                                                      'obligatorio': requisito.obligatorio})
                    return render(request, "adm_convocatoriaexamensuficiencia/editrequisitoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'delrequisitoexamen':
                try:
                    data['title'] = u'Eliminar requisito consultorio'
                    data['requisito'] = RequisitosDetalleConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delrequisitoexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'addlistado':
                try:
                    data['title'] = u'Adicionar archivo'
                    data['proceso'] = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = ArchivoListadoAprobadosExamenComplexivoForm()
                    return render(request, "adm_convocatoriaexamensuficiencia/addlistado.html", data)
                except Exception as ex:
                    pass

            if action == 'delregistro':
                try:
                    data['title'] = u'Eliminar registro'
                    data['registro'] = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delregistro.html", data)
                except Exception as ex:
                    pass

            if action == 'delregistroubi':
                try:
                    data['title'] = u'Eliminar registro'
                    data['registro'] = ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delregistroubi.html", data)
                except Exception as ex:
                    pass

            if action == 'tablasponderacion':

                try:
                    data['title'] = u'Tablas de Ponderación'
                    data['rangosubicacion'] = RangosNotasExamenUbicacionIngles.objects.all().order_by('nivel')
                    data['rangosvalidacion'] = RangosNotasExamenValidacionIngles.objects.all().order_by('nota')
                    data['carrera'] = True
                    return render(request, "adm_convocatoriaexamensuficiencia/tablasponderacion.html", data)
                except Exception as ex:
                    return HttpResponseServerError(f"Error inesperado: {str(ex)}")

            if action == 'addrangoubicacion':
                try:
                    data['title'] = u'Adicionar rango de puntaje para examen de Ubicación'
                    form = RangosNotasExamenInglesForm()
                    data['form'] = form
                    return render(request, "adm_convocatoriaexamensuficiencia/addrangoubicacion.html", data)
                except Exception as ex:
                    return HttpResponseServerError(f"Error inesperado: {str(ex)}")

            if action == 'editrangonota':
                try:
                    data['title'] = u'Editar rango de puntaje para examen de Ubicación'
                    data['rango'] = rango = RangosNotasExamenUbicacionIngles.objects.get(pk=request.GET['id'])
                    data['form'] = form = RangosNotasExamenInglesForm(initial={'inicio': rango.inicio,
                                                                               'fin': rango.fin,
                                                                               'nivel': rango.nivel})
                    return render(request, "adm_convocatoriaexamensuficiencia/editrangoubicacion.html", data)
                except Exception as ex:
                    return HttpResponseServerError(f"Error inesperado: {str(ex)}")

            if action == 'delrango':
                try:
                    data['title'] = u'Eliminar rango de puntaje'
                    data['rango'] = rango = RangosNotasExamenUbicacionIngles.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delrango.html", data)
                except Exception as ex:
                    pass

            if action == 'addrangovalidacion':
                try:
                    data['title'] = u'Adicionar rango para la examen de Validación'
                    form = RangosNotasExamenInglesForm()
                    data['form'] = form
                    return render(request, "adm_convocatoriaexamensuficiencia/addrangovalidacion.html", data)
                except Exception as ex:
                    return HttpResponseServerError(f"Error inesperado: {str(ex)}")

            if action == 'editrangonotaval':
                try:
                    data['title'] = u'Editar rango de puntaje para examen de Validación'
                    data['rango'] = rango = RangosNotasExamenValidacionIngles.objects.get(pk=request.GET['id'])
                    data['form'] = form = RangosNotasExamenInglesForm(initial={'inicio': rango.inicio,
                                                                               'fin': rango.fin,
                                                                               'aprueba': rango.aprueba,
                                                                               'nota': rango.nota})
                    return render(request, "adm_convocatoriaexamensuficiencia/editrangoubicacionval.html", data)
                except Exception as ex:
                    return HttpResponseServerError(f"Error inesperado: {str(ex)}")

            if action == 'delrangoval':
                try:
                    data['title'] = u'Eliminar rango de puntaje'
                    data['rango'] = rango = RangosNotasExamenValidacionIngles.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/delrangoval.html", data)
                except Exception as ex:
                    pass

            if action == 'subirnotas':
                try:
                    data['title'] = u'Subir Notas Examen Validación'
                    data['form'] = ImportarArchivoXLSPeriodoForm()
                    data['convocatoria'] =ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/subirnotas.html", data)
                except Exception as ex:
                    pass

            if action == 'subirnotasubi':
                try:
                    data['title'] = u'Subir Notas Examen Ubicación'
                    data['form'] = ImportarArchivoXLSPeriodoForm()
                    data['convocatoria'] =ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/subirnotasubi.html", data)
                except Exception as ex:
                    pass

            if action == 'pasarrecord':
                try:
                    data['title'] = u'Pasar Homologaciones al record'
                    data['aplicante'] = proceso= ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/pasarrecord.html", data)
                except Exception as ex:
                    pass

            if action == 'pasarrecordubi':
                try:
                    data['title'] = u'Pasar Homologaciones al record'
                    data['aplicante'] = proceso= ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/pasarrecordubi.html", data)
                except Exception as ex:
                    pass

            if action == 'generarrubrosvalidacion':
                try:
                    data['title'] = u'Generar Rubros de Aprobación de Ingles'
                    data['aplicante'] = proceso= ProcesoAplicanteExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/generarrubrosvalidacion.html", data)
                except Exception as ex:
                    pass

            if action == 'editfechaexamen':
                try:
                    data['title'] = u'Editar fecha Rendición de Prueba'
                    data['proceso'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = FechaExamenSuficienciaForm(initial={'fecha': detalle.fechaexamen})
                    return render(request, "adm_convocatoriaexamensuficiencia/editfechaexamen.html", data)
                except Exception as ex:
                    pass

            if action == 'autoregistrar':
                try:
                    data['title'] = u'Registrar Alumnos de Primer Nivel'
                    data['convocatoria'] = detalle = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    return render(request, "adm_convocatoriaexamensuficiencia/autoregistrar.html", data)
                except Exception as ex:
                    pass

            if action == 'addregistroubi':
                try:
                    data['title'] = u'Adicionar registro'
                    data['convocatoria'] = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = RegistroEstudianteExamenInglesForm()
                    return render(request, "adm_convocatoriaexamensuficiencia/addregistroubi.html", data)
                except Exception as ex:
                    pass

            if action == 'addregistroval':
                try:
                    data['title'] = u'Adicionar registro'
                    data['convocatoria'] = ConvocatoriaExamenSuficiencia.objects.get(pk=request.GET['id'])
                    data['form'] = RegistroEstudianteExamenInglesForm()
                    return render(request, "adm_convocatoriaexamensuficiencia/addregistroval.html", data)
                except Exception as ex:
                    pass

            return url_back(request, ex=ex if 'ex' in locals() else None)
        else:
            try:
                data['title'] = u'Procesos de convocatoria al examen de suficiencia'
                search = None
                ids = None
                if 's' in request.GET:
                    search = request.GET['s'].strip()
                    procesos = ConvocatoriaExamenSuficiencia.objects.filter(nombre__icontains=search,
                                                                            periodo=request.session['periodo'],
                                                                            sede=sedeseleccionada).exclude(tipoconvocatoria=tipo_convocatoria_b1()).distinct()
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    procesos = ConvocatoriaExamenSuficiencia.objects.filter(id=ids, sede=sedeseleccionada,
                                                                            periodo=request.session['periodo']).exclude(tipoconvocatoria=tipo_convocatoria_b1())
                else:
                    procesos = ConvocatoriaExamenSuficiencia.objects.filter(sede=sedeseleccionada,
                                                                            periodo=request.session['periodo']).exclude(tipoconvocatoria=tipo_convocatoria_b1())
                paging = MiPaginador(procesos, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session and 'paginador_url' in request.session:
                        if request.session['paginador_url'] == 'adm_convocatoriaexamensuficiencia':
                            paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    page = paging.page(p)
                except:
                    p = 1
                    page = paging.page(p)
                request.session['paginador'] = p
                request.session['paginador_url'] = 'adm_convocatoriaexamensuficiencia'
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['procesos'] = page.object_list
                data['reporte_0'] = obtener_reporte('inscritosexamenidiomas')
                return render(request, "adm_convocatoriaexamensuficiencia/view.html", data)
            except Exception as ex:
                return HttpResponseRedirect('/')
