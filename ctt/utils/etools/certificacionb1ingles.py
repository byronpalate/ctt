# coding=utf-8
from io import BytesIO
from datetime import datetime

import openpyxl
from django import forms
from django.http import HttpResponse

from settings import RUBRO_OTRO_EXAMEN_UBICACION_ID, TIPO_IVA_0_ID
from ctt.forms import (
    BaseForm,
    DetalleConvocatoriaExamenSuficienciaForm,
    FechaExamenSuficienciaForm,
    ImportarArchivoXLSPeriodoForm,
    ProcesoExamenSuficienciaForm,
    RequisitosDetalleFechaProcesoExamenSuficienciaForm,
    RegistroEstudianteExamenInglesForm
)
from ctt.funciones import MiPaginador, bad_json, generar_nombre
from ctt.models import (
    ConvocatoriaExamenSuficiencia,
    DetalleConvocatoriaExamenSuficiencia,
    Inscripcion,
    ProcesoAplicanteExamenSuficiencia,
    RequisitosDetalleConvocatoriaExamenSuficiencia,
    Rubro,
    RubroOtro,
    TipoConvocatoriaExamenSuficiencia,
    null_to_numeric
)


TIPO_CONVOCATORIA_B1_ID = 4
TIPO_CONVOCATORIA_B1_RUBRO = 'CERTIFICACION B1'
NOTA_MINIMA_CERTIFICACION_B1 = 7
VALOR_EXAMEN_CERTIFICACION_B1 = 50.00

# IDs usados actualmente por el SGA para los niveles/modulos de Ingles.
INGLES_ASIGNATURAS_IDS = (4311, 4312, 4313, 4316, 4317, 4318, 4319, 4320)
INGLES_NIVELES_REQUERIDOS = 6


class NotaCertificacionB1Form(BaseForm):
    notaexamen = forms.FloatField(
        label=u'Nota examen certificacion B1',
        required=True,
        initial='0.00',
        widget=forms.TextInput(attrs={
            'class': 'imp-numbermed-center',
            'decimales': '2'
        })
    )

    def extra_paramaters(self):
        self.fields['formbase'].initial = 'ajaxformdinamicbs.html'
        self.fields['formwidth'].initial = 'md'


class ArchivoCertificacionB1Form(BaseForm):
    archivo = forms.FileField(
        label=u'Seleccione certificado',
        required=True,
        help_text=u'Tamano maximo permitido 10Mb, en formato pdf, jpg o png',
        widget=forms.FileInput(attrs={
            'accept': '.pdf,.jpg,.png',
            'data-max-file-size': 10485760
        })
    )

    def extra_paramaters(self):
        self.fields['formbase'].initial = 'ajaxformdinamicbs.html'
        self.fields['formwidth'].initial = 'md'


def paginar(request, queryset, modulo, per_page=25):
    paging = MiPaginador(queryset, per_page)
    p = 1

    try:
        paginasesion = 1
        if 'paginador' in request.session and 'paginador_url' in request.session:
            if request.session['paginador_url'] == modulo:
                paginasesion = int(request.session['paginador'])

        p = int(request.GET['page']) if 'page' in request.GET else paginasesion
        page = paging.page(p)
    except Exception:
        p = 1
        page = paging.page(p)

    request.session['paginador'] = p
    request.session['paginador_url'] = modulo

    return paging, page, p


def tipo_convocatoria_b1():
    tipo = TipoConvocatoriaExamenSuficiencia.objects.filter(
        pk=TIPO_CONVOCATORIA_B1_ID
    ).first()

    if not tipo:
        raise ValueError(
            u'No existe el tipo de convocatoria B1 con ID %s.' %
            TIPO_CONVOCATORIA_B1_ID
        )

    return tipo


def convocatorias_b1():
    return ConvocatoriaExamenSuficiencia.objects.filter(
        tipoconvocatoria_id=TIPO_CONVOCATORIA_B1_ID
    )


def get_convocatoria_b1(pk):
    return convocatorias_b1().get(pk=pk)


def get_detalle_b1(pk):
    return DetalleConvocatoriaExamenSuficiencia.objects.get(
        pk=pk,
        convocatoriaconsultorio__tipoconvocatoria_id=TIPO_CONVOCATORIA_B1_ID
    )


def get_registro_b1(pk):
    return ProcesoAplicanteExamenSuficiencia.objects.get(
        pk=pk,
        convocatoria__tipoconvocatoria_id=TIPO_CONVOCATORIA_B1_ID
    )


def get_requisito_b1(pk):
    return RequisitosDetalleConvocatoriaExamenSuficiencia.objects.get(
        pk=pk,
        detalleproceso__convocatoriaconsultorio__tipoconvocatoria_id=TIPO_CONVOCATORIA_B1_ID
    )


def preparar_form_proceso_b1(form):
    form.fields['tipoconvocatoria'].queryset = TipoConvocatoriaExamenSuficiencia.objects.filter(
        pk=TIPO_CONVOCATORIA_B1_ID
    )
    form.fields['tipoconvocatoria'].initial = TIPO_CONVOCATORIA_B1_ID
    form.fields['tipoconvocatoria'].widget = forms.HiddenInput()
    return form


def proceso_form_inicial_b1(proceso=None):
    if proceso:
        initial = {
            'nombre': proceso.nombre,
            'tipoconvocatoria': proceso.tipoconvocatoria,
            'modalidad': proceso.modalidad.all(),
            'coordinacion': proceso.coordinacion.all(),
            'fechainicio': proceso.fechainicio,
            'fechafin': proceso.fechafin,
            'autoregistro': proceso.autoregistro,
            'fechainicioautoregistro': proceso.fechainicioautoregistro,
            'fechafinautoregistro': proceso.fechafinautoregistro,
            'mensaje': proceso.mensaje,
        }
    else:
        initial = {
            'tipoconvocatoria': tipo_convocatoria_b1(),
            'fechainicio': datetime.now().date(),
            'fechafin': datetime.now().date()
        }

    return preparar_form_proceso_b1(
        ProcesoExamenSuficienciaForm(initial=initial)
    )


def guardar_proceso_b1(request, proceso=None):
    form = ProcesoExamenSuficienciaForm(request.POST)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    if not form.cleaned_data['modalidad']:
        return False, bad_json(mensaje=u'Debe escoger al menos una modalidad.')

    if not form.cleaned_data['coordinacion']:
        return False, bad_json(mensaje=u'Debe escoger al menos una coordinacion.')

    sedeseleccionada = request.session['coordinacionseleccionada'].sede

    if proceso is None:
        proceso = ConvocatoriaExamenSuficiencia(
            periodo=request.session['periodo'],
            sede=sedeseleccionada
        )

    proceso.nombre = form.cleaned_data['nombre']
    proceso.tipoconvocatoria_id = TIPO_CONVOCATORIA_B1_ID
    proceso.fechainicio = form.cleaned_data['fechainicio']
    proceso.fechafin = form.cleaned_data['fechafin']
    proceso.mensaje = form.cleaned_data['mensaje']
    proceso.autoregistro = form.cleaned_data['autoregistro']
    proceso.fechainicioautoregistro = form.cleaned_data['fechainicioautoregistro']
    proceso.fechafinautoregistro = form.cleaned_data['fechafinautoregistro']

    if proceso.pk is None:
        proceso.activo = True if form.cleaned_data['autoregistro'] else False

    proceso.save(request)
    proceso.modalidad.set(form.cleaned_data['modalidad'])
    proceso.coordinacion.set(form.cleaned_data['coordinacion'])

    return True, proceso


def bloqueo_convocatoria_cerrada(proceso):
    if getattr(proceso, 'cerrado', False):
        return u'La convocatoria se encuentra cerrada. No se permiten modificaciones.'
    return None


def fecha_examen_form_b1(proceso):
    return FechaExamenSuficienciaForm(initial={
        'fecha': proceso.fechaexamen
    })


def guardar_fecha_examen_b1(request, proceso):
    form = FechaExamenSuficienciaForm(request.POST)

    if form.is_valid():
        proceso.fechaexamen = form.cleaned_data['fecha']
        proceso.save(request)
        return True, proceso

    return False, bad_json(error=6, form=form)


def detalle_form_inicial_b1(detalle=None):
    if detalle:
        return DetalleConvocatoriaExamenSuficienciaForm(initial={
            'descripcion': detalle.descripcion,
            'inicio': detalle.inicio,
            'fin': detalle.fin,
            'informativo': detalle.informativo
        })

    return DetalleConvocatoriaExamenSuficienciaForm()


def guardar_detalle_b1(request, proceso=None, detalle=None):
    form = DetalleConvocatoriaExamenSuficienciaForm(request.POST)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    if form.cleaned_data['fin'] < form.cleaned_data['inicio']:
        return False, bad_json(mensaje=u'La fecha fin no puede ser menor a la fecha inicio.')

    if detalle is None:
        detalle = DetalleConvocatoriaExamenSuficiencia(
            convocatoriaconsultorio=proceso
        )

    detalle.descripcion = form.cleaned_data['descripcion']
    detalle.inicio = form.cleaned_data['inicio']
    detalle.fin = form.cleaned_data['fin']
    detalle.informativo = form.cleaned_data['informativo']
    detalle.save(request)

    return True, detalle


def requisito_form_inicial_b1(requisito=None):
    if requisito:
        return RequisitosDetalleFechaProcesoExamenSuficienciaForm(initial={
            'tipo': requisito.tipo,
            'obligatorio': requisito.obligatorio
        })

    return RequisitosDetalleFechaProcesoExamenSuficienciaForm()


def guardar_requisito_b1(request, detalle=None, requisito=None):
    form = RequisitosDetalleFechaProcesoExamenSuficienciaForm(request.POST)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    if requisito is None:
        requisito = RequisitosDetalleConvocatoriaExamenSuficiencia(
            detalleproceso=detalle
        )

    requisito.tipo = form.cleaned_data['tipo']
    requisito.obligatorio = form.cleaned_data['obligatorio']
    requisito.save(request)

    return True, requisito


def registro_estudiante_form_b1():
    return RegistroEstudianteExamenInglesForm()


def registrar_estudiante_b1(request, proceso):
    form = RegistroEstudianteExamenInglesForm(request.POST)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    inscripcion = Inscripcion.objects.get(pk=form.cleaned_data['inscripcion'])

    mensaje = validar_postulacion_b1(inscripcion, proceso)
    if mensaje:
        return False, bad_json(mensaje=mensaje)

    if proceso.registro(inscripcion):
        return False, bad_json(
            mensaje=u'El estudiante ya se encuentra registrado en esta convocatoria.'
        )

    registro = crear_registro_b1(proceso, inscripcion, request)

    return True, registro


def nota_certificacion_form_b1(registro):
    return NotaCertificacionB1Form(initial={
        'notaexamen': registro.notaexamenconocimientos
    })


def calificar_registro_b1(request, registro):
    if not registro.pagada():
        return False, bad_json(
            mensaje=u'No se puede registrar la nota porque el examen no esta pagado.'
        )

    form = NotaCertificacionB1Form(request.POST)

    if form.is_valid():
        calcular_resultado_b1(registro, form.cleaned_data['notaexamen'], request)
        return True, registro

    return False, bad_json(error=6, form=form)


def importar_notas_form_b1():
    return ImportarArchivoXLSPeriodoForm()


def certificado_form_b1():
    return ArchivoCertificacionB1Form()


def procesar_subida_notas_b1(request, proceso):
    form = ImportarArchivoXLSPeriodoForm(request.POST, request.FILES)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    workbook = openpyxl.load_workbook(request.FILES['archivo'], data_only=True)
    sheet = workbook.worksheets[0]

    registros = {
        r.inscripcion.persona.cedula: r
        for r in proceso.procesoaplicanteexamensuficiencia_set.select_related(
            'inscripcion__persona'
        )
    }

    actualizados = 0
    omitidos = 0

    for linea, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if linea == 1 or not row or len(row) < 4:
            continue

        cedula = str(row[0]).strip() if row[0] else ''
        nota = row[3]

        if cedula in registros and nota is not None:
            if not registros[cedula].pagada():
                omitidos += 1
                continue

            calcular_resultado_b1(registros[cedula], nota, request)
            actualizados += 1

    return True, {
        'actualizados': actualizados,
        'omitidos': omitidos
    }


def procesar_certificado_b1(request, registro):
    form = ArchivoCertificacionB1Form(request.POST, request.FILES)

    if not form.is_valid():
        return False, bad_json(error=6, form=form)

    guardar_certificado_b1(registro, request.FILES['archivo'], request)

    return True, registro


def eliminar_certificado_b1(registro, request):
    if registro.archivo:
        registro.archivo.delete(save=False)
        registro.archivo = None
        registro.save(request)

    return registro


def records_ingles_record(inscripcion):
    orden = {
        asignatura_id: idx
        for idx, asignatura_id in enumerate(INGLES_ASIGNATURAS_IDS)
    }

    records = list(
        inscripcion.recordacademico_set.filter(
            asignatura_id__in=INGLES_ASIGNATURAS_IDS,
            aprobada=True
        ).select_related('asignatura')
    )

    records.sort(
        key=lambda r: (
            orden.get(r.asignatura_id, 999),
            r.fecha or datetime.min.date(),
            r.id
        )
    )

    return records


def records_ingles_aprobados(inscripcion):
    return records_ingles_record(inscripcion)[:INGLES_NIVELES_REQUERIDOS]


def datos_niveles_ingles(inscripcion):
    todos = records_ingles_record(inscripcion)
    records = todos[:INGLES_NIVELES_REQUERIDOS]

    notas = [float(record.nota or 0) for record in records]
    promedio = null_to_numeric(sum(notas) / len(notas), 2) if notas else 0

    return {
        'records': records,
        'todos': todos,
        'ids': INGLES_ASIGNATURAS_IDS,
        'cantidad': len(records),
        'promedio': promedio,
        'aporte60': null_to_numeric(promedio * 0.6, 2),
        'cumple': len(records) >= INGLES_NIVELES_REQUERIDOS,
    }


def detalle_calculo_b1(registro):
    datos = datos_niveles_ingles(registro.inscripcion)

    aporte_examen = null_to_numeric(
        float(registro.notaexamenconocimientos or 0) * 0.4,
        2
    )

    return {
        'registro': registro,
        'datosniveles': datos,
        'aporte_niveles': datos['aporte60'],
        'aporte_examen': aporte_examen,
        'notafinal': null_to_numeric(datos['aporte60'] + aporte_examen, 2),
    }


def carrera_obligatoria_certificacion_b1(inscripcion):
    nombre = (
        inscripcion.carrera.nombre
        if inscripcion and inscripcion.carrera
        else ''
    ).lower()

    return 'medicina' in nombre or 'enfermer' in nombre


def validar_postulacion_b1(inscripcion, proceso):
    if getattr(proceso, 'cerrado', False):
        return u'La convocatoria se encuentra cerrada.'

    if proceso.coordinacion.exists() and not proceso.coordinacion.filter(
        id=inscripcion.coordinacion_id
    ).exists():
        return u'Usted no pertenece a la facultad designada para esta convocatoria.'

    if proceso.modalidad.exists() and not proceso.modalidad.filter(
        id=inscripcion.modalidad_id
    ).exists():
        return u'Usted no pertenece a la modalidad designada para esta convocatoria.'

    datos = datos_niveles_ingles(inscripcion)

    if not datos['cumple']:
        return u'Para aplicar a certificacion B1 tiene que tener aprobados 6 niveles de Ingles.'

    return None


def nombre_rubro_certificacion_b1(proceso):
    nombrerubro = ''

    if proceso and proceso.tipoconvocatoria:
        nombrerubro = proceso.tipoconvocatoria.nombrerubro or ''

    nombrerubro = nombrerubro.strip() if nombrerubro else TIPO_CONVOCATORIA_B1_RUBRO

    return u'EXAMEN DE %s DE INGLES' % nombrerubro


def crear_rubro_certificacion_b1(registro, request):
    rubro_actual = registro.rubropagar()

    if rubro_actual:
        return rubro_actual

    hoy = datetime.now().date()

    rubro = Rubro(
        inscripcion=registro.inscripcion,
        valor=VALOR_EXAMEN_CERTIFICACION_B1,
        iva_id=TIPO_IVA_0_ID,
        valortotal=VALOR_EXAMEN_CERTIFICACION_B1,
        saldo=VALOR_EXAMEN_CERTIFICACION_B1,
        periodo=registro.convocatoria.periodo,
        fecha=hoy,
        fechavence=hoy
    )
    rubro.save(request)

    rubrootro = RubroOtro(
        rubro=rubro,
        tipo_id=RUBRO_OTRO_EXAMEN_UBICACION_ID
    )
    rubrootro.save(request)

    rubro.actulizar_nombre(
        nombre_rubro_certificacion_b1(registro.convocatoria)
    )

    return rubro


def crear_registro_b1(proceso, inscripcion, request):
    datos = datos_niveles_ingles(inscripcion)

    registro = ProcesoAplicanteExamenSuficiencia(
        convocatoria=proceso,
        inscripcion=inscripcion,
        fechaaplicacion=datetime.today(),
        promedionotas=datos['promedio'],
        fechatope=datetime.now().date()
    )
    registro.save(request)

    crear_rubro_certificacion_b1(registro, request)

    return registro


def calcular_resultado_b1(registro, notaexamen, request=None):
    if isinstance(notaexamen, str):
        notaexamen = notaexamen.strip().replace(',', '.')

    notaexamen = null_to_numeric(float(notaexamen or 0), 2)

    if notaexamen < 0 or notaexamen > 10:
        raise ValueError(u'La nota del examen debe estar entre 0 y 10.')

    datos = datos_niveles_ingles(registro.inscripcion)

    if not datos['cumple']:
        raise ValueError(u'El estudiante no registra 6 niveles de Ingles aprobados.')

    notafinal = null_to_numeric(
        (datos['promedio'] * 0.6) + (notaexamen * 0.4),
        2
    )

    examen_aprobado = notaexamen >= NOTA_MINIMA_CERTIFICACION_B1
    final_aprobado = notafinal >= NOTA_MINIMA_CERTIFICACION_B1

    registro.promedionotas = datos['promedio']
    registro.notaexamenconocimientos = notaexamen
    registro.notafinal = notafinal
    registro.aproboexamenconocimientos = examen_aprobado
    registro.aprobado = examen_aprobado and final_aprobado
    registro.cerrada = True

    if request:
        registro.save(request)
    else:
        registro.save()

    return notafinal


def guardar_certificado_b1(registro, archivo, request):
    if not registro.aprobado:
        raise ValueError(
            u'Solo se puede cargar certificado cuando la nota final es aprobatoria.'
        )

    archivo._name = generar_nombre('certificadob1ingles_', archivo._name)
    registro.archivo = archivo
    registro.save(request)

    return registro.archivo


def descargar_registrados_b1(proceso):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registrados'

    ws.append([
        'CEDULA',
        'ESTUDIANTE',
        'INSCRIPCION',
        'NOTA_EXAMEN',
        'PAGADO',
        'PROMEDIO_NIVELES',
        'APORTE_60',
        'APORTE_40',
        'NOTA_FINAL',
        'ESTADO',
        'OBSERVACION'
    ])

    registrados = proceso.procesoaplicanteexamensuficiencia_set.select_related(
        'inscripcion__persona',
        'inscripcion__carrera',
        'inscripcion__modalidad',
        'inscripcion__sede'
    ).order_by('inscripcion__persona__apellido1')

    for registro in registrados:
        calculo = detalle_calculo_b1(registro)
        pagado = registro.pagada()

        estado = 'PENDIENTE'
        if registro.cerrada:
            estado = 'APROBADO' if registro.aprobado else 'NO APROBADO'

        ws.append([
            registro.inscripcion.persona.cedula,
            registro.inscripcion.persona.nombre_completo(),
            registro.inscripcion_id,
            registro.notaexamenconocimientos or 0,
            'SI' if pagado else 'NO',
            calculo['datosniveles']['promedio'],
            calculo['aporte_niveles'],
            calculo['aporte_examen'],
            calculo['notafinal'],
            estado,
            '' if pagado else 'No se permite cargar nota hasta registrar pago',
        ])

    for column in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
        ws.column_dimensions[column].width = 18

    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['K'].width = 45

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="registrados_b1_%s.xlsx"' % proceso.id

    return response
